"""
CSV and Excel Renderers for Multi-Format Artifact Engine (v2.5.1 - FEAT-ARTIFACT-002)

Provides renderers for tabular data export:
- CsvRenderer: Exports data to CSV format
- ExcelRenderer: Exports data to XLSX format with formatting

Used primarily for:
- Computo Metrico (Bill of Quantities)
- Data exports
- Tabular reports

Architecture:
- AI extracts structured data as JSON (json_extraction mode)
- Python performs calculations, formatting, and file generation

Usage:
    from renderers.csv_renderer import CsvRenderer, ExcelRenderer
    from renderers.base import RenderContext, OutputFormat

    context = RenderContext(
        title="Computo Metrico",
        content="...",
        extracted_data=[{"item": "X", "qty": 10}],
        format_settings={"freeze_header": True}
    )

    renderer = ExcelRenderer()
    result = renderer.render(context)
"""

import io
import csv
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional, Union
from dataclasses import dataclass

from .base import (
    FormatRenderer,
    OutputFormat,
    RenderContext,
    RenderResult,
    DataExtractor,
    TableDataProcessor,
    PostProcessType,
)

logger = logging.getLogger(__name__)


@dataclass
class ColumnDefinition:
    """Definition for a table column."""

    name: str
    type: str = "string"  # string, number, currency, formula, auto_increment
    width: int = 15
    decimals: int = 2
    source: Optional[str] = None
    formula: Optional[str] = None
    format: Optional[str] = None


class CsvRenderer(FormatRenderer):
    """
    Renderer for CSV (Comma-Separated Values) format.

    Features:
    - Configurable delimiter and encoding
    - BOM support for Excel compatibility
    - Header row generation
    - Proper escaping of special characters
    """

    def __init__(self, settings: Optional[Dict[str, Any]] = None):
        """Initialize CSV renderer."""
        super().__init__(settings)
        self._delimiter = self._get_setting("csv_delimiter", ",")
        self._encoding = self._get_setting("csv_encoding", "utf-8-sig")
        self._quote_char = self._get_setting("csv_quote_char", '"')

    @property
    def output_format(self) -> OutputFormat:
        """Return CSV output format."""
        return OutputFormat.CSV

    def render(self, context: RenderContext) -> RenderResult:
        """
        Render tabular data to CSV format.

        Args:
            context: RenderContext with extracted_data or content

        Returns:
            RenderResult with CSV bytes
        """
        import time

        start_time = time.time()
        self._log_render_start(context)

        try:
            # Get data from context
            data = self._get_data(context)
            if not data:
                return RenderResult.failure(
                    format=self.output_format,
                    error="No data available for CSV export",
                    filename=self.generate_filename(context.title, context.blueprint_id),
                )

            # Get column definitions if available
            columns = self._get_columns(context)

            # Generate CSV content
            csv_bytes = self._generate_csv(data, columns)

            # Create result
            duration_ms = int((time.time() - start_time) * 1000)
            filename = self.generate_filename(context.title, context.blueprint_id)

            result = RenderResult(
                content=csv_bytes,
                format=self.output_format,
                filename=filename,
                render_time_ms=duration_ms,
                metadata={
                    "rows": len(data),
                    "columns": len(columns) if columns else len(data[0]) if data else 0,
                    "encoding": self._encoding,
                },
            )

            self._log_render_complete(result, duration_ms)
            return result

        except Exception as e:
            logger.error(f"[CSV] Render error: {e}", exc_info=True)
            return RenderResult.failure(
                format=self.output_format,
                error=str(e),
                filename=self.generate_filename(context.title, context.blueprint_id),
            )

    def _get_data(self, context: RenderContext) -> List[Dict]:
        """Extract data from context."""
        # First try extracted_data
        if context.extracted_data:
            if isinstance(context.extracted_data, list):
                return context.extracted_data
            elif isinstance(context.extracted_data, dict):
                # Handle {"columns": [...], "rows": [...]} format
                if "rows" in context.extracted_data:
                    columns = context.extracted_data.get("columns", [])
                    rows = context.extracted_data.get("rows", [])
                    return [dict(zip(columns, row)) for row in rows]
                return [context.extracted_data]

        # Try extracting from content
        extracted = DataExtractor.extract_json(context.content)
        if extracted:
            if isinstance(extracted, list):
                return extracted
            elif isinstance(extracted, dict):
                if "rows" in extracted:
                    columns = extracted.get("columns", [])
                    rows = extracted.get("rows", [])
                    return [dict(zip(columns, row)) for row in rows]
                return [extracted]

        return []

    def _get_columns(self, context: RenderContext) -> List[ColumnDefinition]:
        """Get column definitions from context."""
        columns_config = context.format_settings.get("columns", [])
        columns = []

        for col_config in columns_config:
            col = ColumnDefinition(
                name=col_config.get("name", ""),
                type=col_config.get("type", "string"),
                width=col_config.get("width", 15),
                decimals=col_config.get("decimals", 2),
                source=col_config.get("source"),
                formula=col_config.get("formula"),
                format=col_config.get("format"),
            )
            columns.append(col)

        return columns

    def _generate_csv(
        self,
        data: List[Dict],
        columns: List[ColumnDefinition],
    ) -> bytes:
        """Generate CSV content from data."""
        output = io.StringIO()

        # Determine column names
        if columns:
            fieldnames = [col.name for col in columns if col.type != "auto_increment" and col.type != "formula"]
        elif data:
            fieldnames = list(data[0].keys())
        else:
            fieldnames = []

        writer = csv.DictWriter(
            output,
            fieldnames=fieldnames,
            delimiter=self._delimiter,
            quotechar=self._quote_char,
            quoting=csv.QUOTE_MINIMAL,
            extrasaction='ignore',
        )

        # Write header
        if self._get_setting("include_header", True):
            writer.writeheader()

        # Write data rows
        for row in data:
            # Map source columns to output columns if needed
            output_row = {}
            if columns:
                for col in columns:
                    if col.type in ("auto_increment", "formula"):
                        continue
                    source_key = col.source or col.name
                    value = row.get(source_key, row.get(col.name, ""))
                    output_row[col.name] = self._format_value(value, col)
            else:
                output_row = row

            writer.writerow(output_row)

        # Encode with specified encoding
        csv_content = output.getvalue()
        return csv_content.encode(self._encoding)

    def _format_value(self, value: Any, column: ColumnDefinition) -> str:
        """Format a value according to column type."""
        if value is None:
            return ""

        if column.type == "number":
            try:
                num = float(value)
                return f"{num:.{column.decimals}f}"
            except (ValueError, TypeError):
                return str(value)

        if column.type == "currency":
            try:
                num = float(value)
                return f"{num:.{column.decimals}f}"
            except (ValueError, TypeError):
                return str(value)

        return str(value)


class ExcelRenderer(FormatRenderer):
    """
    Renderer for Excel XLSX format.

    Features:
    - Multiple sheets support
    - Column auto-width
    - Header freeze
    - Cell formatting (number, currency, date)
    - Formula support
    - Auto-filter
    - Totals row
    - Conditional formatting

    Used primarily for Computo Metrico with price list integration.
    """

    def __init__(self, settings: Optional[Dict[str, Any]] = None):
        """Initialize Excel renderer."""
        super().__init__(settings)
        self._sheet_name = self._get_setting("xlsx_default_sheet_name", "Data")
        self._auto_width = self._get_setting("xlsx_auto_column_width", True)
        self._freeze_header = self._get_setting("xlsx_freeze_header", True)

    @property
    def output_format(self) -> OutputFormat:
        """Return XLSX output format."""
        return OutputFormat.XLSX

    def render(self, context: RenderContext) -> RenderResult:
        """
        Render tabular data to XLSX format.

        Args:
            context: RenderContext with extracted_data or content

        Returns:
            RenderResult with XLSX bytes
        """
        import time

        start_time = time.time()
        self._log_render_start(context)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("[XLSX] openpyxl not installed")
            return RenderResult.failure(
                format=self.output_format,
                error="openpyxl is required for Excel export",
                filename="",
            )

        try:
            # Get data from context
            data = self._get_data(context)

            # v2.5.1: Handle empty data gracefully - generate Excel with warning
            if not data:
                logger.warning("[XLSX] No data available, generating placeholder Excel")
                data = [{
                    "nota": "Nessun dato strutturato estratto dal report",
                    "descrizione": "Il report è stato generato ma non contiene dati tabellari",
                    "azione": "Verificare il contenuto del report e riprovare",
                }]

            # Get column definitions
            columns = self._get_columns(context)

            # Apply post-processing if needed
            if context.post_process == PostProcessType.APPLY_PRICE_LIST:
                data = self._apply_price_list(data, context)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = context.format_settings.get("sheet_name", self._sheet_name)

            # Add title row if specified
            if context.format_settings.get("include_title", False):
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns) or 7)
                title_cell = ws.cell(row=1, column=1, value=context.title)
                title_cell.font = Font(size=14, bold=True)
                title_cell.alignment = Alignment(horizontal='center')
                start_row = 3
            else:
                start_row = 1

            # Write headers
            header_row = start_row
            if columns:
                for col_idx, col in enumerate(columns, 1):
                    cell = ws.cell(row=header_row, column=col_idx, value=col.name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        bottom=Side(style='thin'),
                        top=Side(style='thin'),
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                    )
            elif data:
                for col_idx, key in enumerate(data[0].keys(), 1):
                    cell = ws.cell(row=header_row, column=col_idx, value=key)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

            # Write data rows
            auto_increment = 1
            for row_idx, row_data in enumerate(data, header_row + 1):
                if columns:
                    for col_idx, col in enumerate(columns, 1):
                        if col.type == "auto_increment":
                            value = auto_increment
                            auto_increment += 1
                        elif col.type == "formula":
                            # Replace {row} placeholder with actual row number
                            formula = col.formula.replace("{row}", str(row_idx))
                            value = formula
                        else:
                            source_key = col.source or col.name
                            value = row_data.get(source_key, row_data.get(col.name, ""))

                        cell = ws.cell(row=row_idx, column=col_idx, value=value)

                        # Apply formatting based on type
                        if col.type in ("number", "currency"):
                            if col.type == "currency":
                                cell.number_format = f'#,##0.{"0" * col.decimals}'
                            else:
                                cell.number_format = f'0.{"0" * col.decimals}'
                        elif col.format == "currency":
                            cell.number_format = '#,##0.00'

                        cell.border = Border(
                            bottom=Side(style='thin'),
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                        )
                else:
                    for col_idx, value in enumerate(row_data.values(), 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

            # Add totals row if specified
            if context.format_settings.get("include_totals", False) and columns:
                totals_row = header_row + len(data) + 1
                for col_idx, col in enumerate(columns, 1):
                    if col.type in ("number", "currency") or col.format == "currency":
                        # Add SUM formula
                        col_letter = get_column_letter(col_idx)
                        formula = f"=SUM({col_letter}{header_row + 1}:{col_letter}{totals_row - 1})"
                        cell = ws.cell(row=totals_row, column=col_idx, value=formula)
                        cell.font = Font(bold=True)
                        cell.number_format = '#,##0.00'
                    elif col_idx == 1:
                        cell = ws.cell(row=totals_row, column=col_idx, value="TOTALE")
                        cell.font = Font(bold=True)

            # Auto-width columns
            if self._auto_width:
                for col_idx, col in enumerate(columns or [{}] * (len(data[0]) if data else 0), 1):
                    width = getattr(col, 'width', None) or 15
                    ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Freeze header row
            if self._freeze_header:
                ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

            # Add auto-filter
            if context.format_settings.get("auto_filter", False) and data:
                last_col = get_column_letter(len(columns) if columns else len(data[0]))
                last_row = header_row + len(data)
                ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"

            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            xlsx_bytes = buffer.getvalue()

            # Create result
            duration_ms = int((time.time() - start_time) * 1000)
            filename = self.generate_filename(context.title, context.blueprint_id)

            result = RenderResult(
                content=xlsx_bytes,
                format=self.output_format,
                filename=filename,
                render_time_ms=duration_ms,
                metadata={
                    "rows": len(data),
                    "columns": len(columns) if columns else len(data[0]) if data else 0,
                    "sheet_name": ws.title,
                    "has_totals": context.format_settings.get("include_totals", False),
                },
            )

            self._log_render_complete(result, duration_ms)
            return result

        except Exception as e:
            logger.error(f"[XLSX] Render error: {e}", exc_info=True)
            return RenderResult.failure(
                format=self.output_format,
                error=str(e),
                filename=self.generate_filename(context.title, context.blueprint_id),
            )

    def _get_data(self, context: RenderContext) -> List[Dict]:
        """Extract data from context."""
        # First try extracted_data
        if context.extracted_data:
            if isinstance(context.extracted_data, list):
                return context.extracted_data
            elif isinstance(context.extracted_data, dict):
                if "rows" in context.extracted_data:
                    columns = context.extracted_data.get("columns", [])
                    rows = context.extracted_data.get("rows", [])
                    return [dict(zip(columns, row)) for row in rows]
                return [context.extracted_data]

        # Try extracting from content
        extracted = DataExtractor.extract_json(context.content)
        if extracted:
            if isinstance(extracted, list):
                return extracted
            elif isinstance(extracted, dict):
                if "rows" in extracted:
                    columns = extracted.get("columns", [])
                    rows = extracted.get("rows", [])
                    return [dict(zip(columns, row)) for row in rows]
                return [extracted]

        return []

    def _get_columns(self, context: RenderContext) -> List[ColumnDefinition]:
        """Get column definitions from context."""
        columns_config = context.format_settings.get("columns", [])
        columns = []

        for col_config in columns_config:
            col = ColumnDefinition(
                name=col_config.get("name", ""),
                type=col_config.get("type", "string"),
                width=col_config.get("width", 15),
                decimals=col_config.get("decimals", 2),
                source=col_config.get("source"),
                formula=col_config.get("formula"),
                format=col_config.get("format"),
            )
            columns.append(col)

        return columns

    def _apply_price_list(
        self,
        data: List[Dict],
        context: RenderContext,
    ) -> List[Dict]:
        """
        Apply price list to extracted quantity data.

        This is the Python post-processing step that applies pricing
        from a configured price list file to the AI-extracted quantities.

        Args:
            data: Extracted data with quantities
            context: Render context with settings

        Returns:
            Data with unit prices applied
        """
        price_list_path = self._get_setting("price_list_path", "")
        if not price_list_path:
            logger.info("[XLSX] No price list configured, skipping price application")
            return data

        # Load price list
        price_lookup = self._load_price_list(price_list_path)
        if not price_lookup:
            logger.warning(f"[XLSX] Could not load price list from: {price_list_path}")
            return data

        # Apply prices
        enriched_data = []
        for row in data:
            enriched_row = dict(row)
            code = row.get("code") or row.get("codice") or ""

            if code and code in price_lookup:
                price_info = price_lookup[code]
                enriched_row["unit_price"] = price_info.get("unit_price", 0.0)
            else:
                enriched_row["unit_price"] = 0.0

            enriched_data.append(enriched_row)

        logger.info(f"[XLSX] Applied prices to {len(enriched_data)} rows")
        return enriched_data

    def _load_price_list(self, path: str) -> Dict[str, Dict]:
        """
        Load price list from file.

        Supports CSV, XLSX, and JSON formats.

        Args:
            path: Path to price list file

        Returns:
            Dictionary mapping codes to price info
        """
        import os
        import json

        if not os.path.exists(path):
            return {}

        ext = os.path.splitext(path)[1].lower()
        price_lookup = {}

        try:
            if ext == ".json":
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    for item in data:
                        code = item.get("code", "")
                        if code:
                            price_lookup[code] = item

            elif ext == ".csv":
                with open(path, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        code = row.get("code", "")
                        if code:
                            price_lookup[code] = {
                                "description": row.get("description", ""),
                                "unit": row.get("unit", ""),
                                "unit_price": float(row.get("unit_price", 0)),
                            }

            elif ext == ".xlsx":
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(path, read_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]

                    code_idx = headers.index("code") if "code" in headers else 0
                    desc_idx = headers.index("description") if "description" in headers else 1
                    unit_idx = headers.index("unit") if "unit" in headers else 2
                    price_idx = headers.index("unit_price") if "unit_price" in headers else 3

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        code = row[code_idx] if code_idx < len(row) else ""
                        if code:
                            price_lookup[str(code)] = {
                                "description": row[desc_idx] if desc_idx < len(row) else "",
                                "unit": row[unit_idx] if unit_idx < len(row) else "",
                                "unit_price": float(row[price_idx]) if price_idx < len(row) and row[price_idx] else 0,
                            }
                    wb.close()
                except ImportError:
                    logger.warning("[XLSX] openpyxl not available for reading price list")

            logger.info(f"[XLSX] Loaded {len(price_lookup)} prices from {path}")

        except Exception as e:
            logger.error(f"[XLSX] Error loading price list: {e}")

        return price_lookup
